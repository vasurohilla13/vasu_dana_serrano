"""
Run this script once to generate the Excel budget tracker files.
Requires: pip install openpyxl
Usage:    python generate_artifacts.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUT_DIR = os.path.join(os.path.dirname(__file__), "environment", "artifacts", "inputs", "files")
os.makedirs(OUT_DIR, exist_ok=True)


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def make_header_fill():
    return PatternFill(fill_type="solid", fgColor="D9E1F2")


# ── Budget Tracker v2 (SIGNAL file — agent should use this) ───────────────

def create_v2():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Val Wedding Expenses"

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = "Val & Marco Wedding — Expense Tracker"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Version and budget row
    ws["A2"] = "Version:"
    ws["B2"] = "v2"
    ws["B2"].font = Font(bold=True)
    ws["C2"] = "Last Updated:"
    ws["D2"] = "2026-10-05"
    ws["E2"] = "Total Budget:"
    ws["F2"] = 1650.00
    ws["F2"].number_format = '"$"#,##0.00'
    ws["F2"].font = Font(bold=True, color="1F5C99")

    # Column headers (row 4)
    headers = ["Date", "Vendor", "Description", "Est. Amount", "Actual Amount (Plaid)", "Status", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2E5FA3")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    # Pre-filled rows (estimated amounts only — agent fills in actuals)
    rows = [
        ("2026-09-22", "The Menger Hotel", "Hotel 2 nights Oct 16-18, San Antonio", 360.00, None, "Pending verification", "Pre-paid"),
        ("2026-10-01", "Flores & Bloom", "Bouquet contribution — split 4 bridesmaids", 50.00, None, "Pending verification", "Venmo"),
        ("2026-10-02", "Elena's Alterations", "Bridesmaid dress alteration", 270.00, None, "Pending verification", ""),
        ("2026-10-05", "Zola Inc", "Wedding registry gift", 160.00, None, "Pending verification", "Over limit — needs approval"),
        ("2026-10-08", "Southwest Airlines", "DEN to SAT roundtrip Oct 16/19", 200.00, None, "Pending verification", ""),
    ]

    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = thin_border()
            if j in (4, 5):
                cell.number_format = '"$"#,##0.00'
            if i % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor="EEF2FA")

    # Totals row
    ws.cell(row=11, column=3, value="ESTIMATED TOTAL").font = Font(bold=True)
    ws.cell(row=11, column=4, value=1040.00).number_format = '"$"#,##0.00'
    ws.cell(row=11, column=4).font = Font(bold=True)

    # Column widths
    widths = [12, 22, 38, 16, 22, 22, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    path = os.path.join(OUT_DIR, "file_05.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ── Budget Tracker v1 (DECOY file — agent must NOT use this) ──────────────

def create_v1():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Val Wedding Expenses"

    ws.merge_cells("A1:G1")
    ws["A1"] = "Val & Marco Wedding — Expense Tracker"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "Version:"
    ws["B2"] = "v1"
    ws["B2"].font = Font(bold=True)
    ws["C2"] = "Last Updated:"
    ws["D2"] = "2026-09-12"
    ws["E2"] = "Total Budget:"
    ws["F2"] = 1800.00   # <-- stale higher budget
    ws["F2"].number_format = '"$"#,##0.00'
    ws["F2"].font = Font(bold=True, color="C00000")

    headers = ["Date", "Vendor", "Description", "Est. Amount", "Actual Amount", "Status", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="767676")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    # Stale estimates from September (all wrong amounts, no actuals)
    rows = [
        ("2026-09-12", "Hotel TBD", "Hotel San Antonio (TBD)", 400.00, None, "Draft", "Checking options"),
        ("2026-09-12", "Florist TBD", "Flowers/bouquet", 60.00, None, "Draft", ""),
        ("2026-09-12", "Alteration shop", "Dress alterations (estimate)", 240.00, None, "Draft", "Quote from Elena's"),
        ("2026-09-12", "Gift — TBD", "Wedding gift", 150.00, None, "Draft", ""),
        ("2026-09-12", "Flights", "Flights DEN-SAT", 220.00, None, "Draft", ""),
        ("2026-09-12", "Incidentals", "Misc", 50.00, None, "Draft", ""),
    ]

    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = thin_border()
            if j in (4, 5):
                cell.number_format = '"$"#,##0.00'

    ws.cell(row=12, column=3, value="ESTIMATED TOTAL").font = Font(bold=True)
    ws.cell(row=12, column=4, value=1120.00).number_format = '"$"#,##0.00'

    widths = [12, 22, 38, 16, 18, 18, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    path = os.path.join(OUT_DIR, "file_06.xlsx")
    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    create_v2()
    create_v1()
    print("\nDone. file_05.xlsx (v2, $1,650 budget) and file_06.xlsx (v1, $1,800 budget) generated.")
